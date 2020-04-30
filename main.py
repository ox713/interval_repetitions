"""
Программа которая должна заменить моё экселевское повторение.
Задачи:

Новая база данных:
    Перевод всего в нужную базу и тогда будет просто импорт
    Обновление базы по мере необходимости
Программа которая бы выводила мои данные в виде html + работа с базой данных:


Чтение:
    Имортировать все нужные файлы которые я добавлю через модуль values
    Анализировать ячейки по дате + смещению, которые тоже будут в отдельном файле
    Выдача нужных ячеек в формате вопрос ответ
    Выдача ячеек на несколько дней
    !Коректное закрытие

Запись
    Запись даты повторений с потверждением(будущим)


Модули
    Проверка везде ли установленна дата
    Проверка везде ли установлен вопрос

    Подгтовка к экспорту в балаболку или скармливание голосовому движку

    Реалзиация систем оценки с помощью андроид приложения и пульта:
    1. Проигрывание и замирание до оценки

Функционал:
    Типы повторений: по дате, по темам, по тегам, по источнику, по языку источника, по качеству усвоения,
    по сложности, по типу ячеек(например код не нужно в аудио), по листам

    Для озвучивания необходимо метить отдельно английские и русские слова
    Практические повторения, то есть берем генерируем пример который нужно запрогать определенным методом и в уме делаем.
Поля:
1.
"""
import datetime
import sqlite3


import config
import mesql
from openpyxl import load_workbook
from datetime import timedelta
from subprocess import Popen, PIPE
from os import listdir

class ParserExcel(object):
    def __init__(self, EXCELS):
        self.excels = EXCELS

    # def import_excel_files(self):
    #     for excel in self.excels:
    #         wb = load_workbook(filename=excel)
    #         self.all_reiteration.append(self.operations.value_operations(wb[wb.sheetnames[0]]))

    def load_excel_file(self, excel):
        # names_of_sheets = []
        excel_file = load_workbook(filename=excel)  # объект загружаеться в память, так что надо работать
        # names_of_sheets.append({excel: wb.sheetnames})

        return excel_file

    def return_data_from_names_of_sheets(self, excel, sheet):
        pass

    def test_print(self, reiteration, sheet):
        for i in reiteration:
            # Вопрос\ответ
            print(sheet.cell(row=i, column=3), 1)
            print(sheet.cell(row=i, column=2), 2)


class OperationsWithParsedExcel(object):

    def __init__(self):
        self.export_data = ExportDoneFile()
        # self.excels = excels
        # self.all_reiteration = all_reiteration

    #
    # def call_import_excel(self, import_class):
    #
    #     pass
    def need_dates(self, needed_date):
        need_dates = []
        today = datetime.date.today()
        date = today + timedelta(days=needed_date)

        for i in config.DATES_PLUS:
            need_dates.append(date - timedelta(days=i))
        return need_dates

    def value_excel_operations(self, sheet):
        # пока так потом может добавить выбор
        if config.TYPE_OF_SELECT == "today":
            # self.export_data.to_txt(, sheet)
            return self.find_reiteration(sheet, 0)

        elif config.TYPE_OF_SELECT == "tomorrow":
            return self.find_reiteration(sheet, 1)

        elif config.TYPE_OF_SELECT == "tag":
            pass

        elif config.TYPE_OF_SELECT == "source":
            pass

        elif config.TYPE_OF_SELECT == "category":
            pass

        elif config.TYPE_OF_SELECT == "language":
            pass

        elif config.TYPE_OF_SELECT == "quality of understand":
            pass

        elif config.TYPE_OF_SELECT == "special code":
            pass

        elif config.TYPE_OF_SELECT == "quality of remembering":
            pass

        elif config.TYPE_OF_SELECT == "list":
            pass

    def end_of_file(self, sheet):
        for i in range(1, 100000):
            if sheet.cell(row=i, column=1).value == "end_of_end":
                print(i, 3)
                return i
        return 10000
    def find_reiteration(self, sheet, needed_date):
        reiteration = []
        end = self.end_of_file(sheet)
        today = datetime.date.today()

        calculate_dates = self.need_dates(needed_date)

        if isinstance(needed_date, datetime.date):
            date = needed_date

        else:  # Если смещение
            date = today + timedelta(days=needed_date)

        if end == None:
            print("FAIL OF END FILE", 4)

        for i in range(1, end):
            #print(sheet.cell(row=i, column=4).value, 5)

            # ПРЕД ищет объетк дата


            if not isinstance(sheet.cell(row=i, column=4).value, datetime.datetime):  # sheet.cell(row=i, column=4).value == None or sheet.cell(row=i, column=4).value =="":
                continue
            # найдя дату перебирает на возможные совпадения с повторениями.

            else:
                cell_value = sheet.cell(row=i, column=4).value
                cell_value = datetime.date(int(str(cell_value)[0:4]), int(str(cell_value)[5:7]), int(str(cell_value)[8:10]))
                for g in calculate_dates:
                    #temp = sheet.cell(row=i, column=4).value + timedelta(days=g) # нужна так как заменил сразу вычеслинными датами
                    if datetime.date(int(str(g)[0:4]), int(str(g)[5:7]), int(str(g)[8:10])) == cell_value:
                        reiteration.append(i)
                        break
        print(reiteration, 6)
        return reiteration


class ExportDoneFile(object):
    def __init__(self):
        pass

    def to_txt(self, sheet, repeat, excel=None, sheet_name=None):

        with open(config.TYPE_OF_SELECT + ".txt", 'a+', encoding="UTF-8") as f:
            counter = 1
            for i in repeat:
                print(sheet.cell(row=i, column=3).value, 61)
                #f.write(f"{excel}\n")
                #f.write(f"{sheet_name}\n")
                f.write(f"Вопрос {counter}\n")
                f.write(str(sheet.cell(row=i, column=3).value) + "\n")

                print(sheet.cell(row=i, column=2).value, 7)
                f.write(f"Ответ {counter}\n")
                f.write(str(sheet.cell(row=i, column=2).value) + "\n")
                #f.write("-----\n\n")
                #counter += 1

    def to_txt_cut_answer_q(self, sheet, repeat, excel=None, sheet_name=None):
        date = str(datetime.date.today())
        counter = 1
        #path = "C:/Users/Ox/Desktop/qa/"
        path = "./qa/"

        for i in repeat:
            with open(path + config.TYPE_OF_SELECT + date + "Вопрос_" + str(counter) + ".txt", 'a+', encoding="UTF-8") as f:
                f.write(f"Вопрос {counter}\n")
                f.write(str(sheet.cell(row=i, column=3).value) + "\n")

            with open(path + config.TYPE_OF_SELECT + date + "Вопрос_" + str(counter) + "о" + ".txt", 'a+', encoding="UTF-8") as f:
                f.write(f"Ответ {counter}\n")
                f.write(str(sheet.cell(row=i, column=2).value) + "\n")
            counter += 1
    def to_base(self, sheet, repeat):
        counter = 1
        database = mesql.SQLighter1("main_base.db")
        for i in repeat:
            tuple_to_base = (
                str(sheet.cell(row=i, column=3).value),
                str(sheet.cell(row=i, column=2).value),
                str(sheet.cell(row=i, column=4).value),
                str(sheet.cell(row=i, column=1).value),
                str(sheet.cell(row=i, column=5).value),
                str(sheet.cell(row=i, column=6).value),
                str(sheet.cell(row=i, column=7).value),
                str(sheet.cell(row=i, column=8).value),
                str(sheet.cell(row=i, column=9).value),
                str(sheet.cell(row=i, column=10).value),
                str(sheet.cell(row=i, column=11).value),
                str(sheet.cell(row=i, column=12).value),
                str(sheet.cell(row=i, column=13).value),
                str(sheet.cell(row=i, column=14).value)

            )
            print(tuple_to_base)


            #connection = sqlite3.connect("main_base", check_same_thread=False)
            #cursor = connection.cursor()
            #with connection:
            database.writing_line("python", tuple_to_base)
            #connection.commit()



class ExportInBalabolku():
    def __init__(self):
        pass

    def export_in_one_file(self, file_name):
        path = "./" + file_name + ".txt"
        # path = "C:/Users/Ox/Desktop/qa/"
        outpath = "C:/Users/Ox/Desktop/qa/" + file_name + ".wav"

        # Голоса катя и прочие требуют админских прав, поэтому надо писать выходной путь полностью
        proc = Popen(
            f"C:/1/balcon.exe -f {path} -w {outpath} -n Katya -enc utf8",#-n aly
            shell=True,
            stdout=PIPE, stderr=PIPE
        )
        proc.wait()  # дождаться выполнения
        res = proc.communicate()  # получить tuple('stdout', 'stderr')
        if proc.returncode:
            print(res[1])
        print('result:', res[0])

    def export_in_many_files(self):
        file_names = listdir(path="./qa/")

        for i in file_names:
            path = 'C:/Base/All/programming/python/interval_repetitions/qa/' + i
            outpath = "C:/Users/Ox/Desktop/qa/" + i[:-4] + ".wav"
            # Голоса катя и прочие требуют админских прав, поэтому надо писать выходной путь полностью
            proc = Popen(
                f"C:/1/balcon.exe -f {path} -w {outpath} -n Katya -enc utf8",#-n aly
                shell=True,
                stdout=PIPE, stderr=PIPE
            )
            proc.wait()  # дождаться выполнения
            res = proc.communicate()  # получить tuple('stdout', 'stderr')
            if proc.returncode:
                print(res[1])

            print('result:', res[0])


    def def_language(self, file_name):
        with open(file_name + ".txt", "r", encoding="utf8") as r:
            with open(file_name + "_done" + ".txt", "w",encoding="utf8") as w:
                buf = None
                for i in r:
                    #i = i.
                    line = i.strip().split(" ")
                    #line = i.split(" ")
                    #if line == None or len(line) < 2: break

                    fin_str = ""
                    for g in line:
                        print(g)
                        # Кириллица
                        # # latin
                        #print(ord(g[0]))
                        # если это аски тогда это английский, исключаем цифры и прочие знаки

                        if 64 < ord(g[0]) < 91 or 96 < ord(g[0]) < 123:
                            if buf == 1:
                                fin_str += g + " "

                            else:
                                new_line = r'<voice required="Language=409"><voice required="Name=VW Julie">' + g + " "
                                fin_str += new_line
                                buf = 1
                        else:

                            if buf == 0:
                                fin_str += g + " "
                            else:
                                new_line = r'<voice required="Language=419"><voice required="Name=Alyona22k">' + g + " "
                                fin_str += new_line
                                buf = 0
                    w.write(fin_str)
                    w.write("\n")


                        # if 97 <= ord(g[0].lower()) <= 122:
                        #     new_line = r'<voice required="Language=409"><voice required="Name=VW Julie">' + g
                        #     w.write(new_line)
                        # # другие надо как-то обозначить
                        # else: pass
                        #
                        # elif 1040 <= ord(g[0])  <= 1103:
                        #     pass


    def cut_on_q_a_parts(self):



        pass


# лучше вынесу все объеты в маин, чтобы логика не путалась, чтобы стал главной базой
def main():
    all_reiteration = []
    operations = OperationsWithParsedExcel()
    loader = ParserExcel(config.EXCELS)
    export_list = ExportDoneFile()
    balabolka = ExportInBalabolku()
    k = 1

    for excel in config.EXCELS:
        print(excel,"####################################################")
        excel_file = loader.load_excel_file(excel)
        for sheet in excel_file.sheetnames:
            print(sheet, "@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@")
            if excel_file[sheet].cell(row=1, column=1).value == "NO":
                continue
            #export_list.to_txt(excel_file[sheet], operations.value_excel_operations(excel_file[sheet]), excel, sheet)
            export_list.to_txt_cut_answer_q(excel_file[sheet], operations.value_excel_operations(excel_file[sheet]), excel, sheet)
            #export_list.to_base(excel_file[sheet], operations.value_excel_operations(excel_file[sheet]))
            print("done", k, 8)
            k += 1
    #balabolka.def_language(config.TYPE_OF_SELECT)
    #balabolka.export_in_one_file(config.TYPE_OF_SELECT)
    balabolka.export_in_many_files()

main()
